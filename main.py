import os
from copy import copy
import sys
import subprocess
import uuid
import shutil
import asyncio
import time
import datetime
import pytz
import numpy as np
import openpyxl
from io import BytesIO
from playwright.async_api import async_playwright
import pandas as pd

LIMIT_QUERY = 499
MAIN_SHEET_NAME = "Data filtrada"
KEYWORDS = ["UNIVERSIDAD", "HOSPITAL", "COLEGIO"]
KEYWORDS_VIDRIOS = ["VENTANA", "MAMPARA", "MURO CORTINA", "VIDRIO"]

DATA_DIR  = os.environ.get("DATA_DIR", "./data")
TMP_DIR   = f"{DATA_DIR}/tmp"
QUERY_DIR = f"{DATA_DIR}/query"
DRIVE_DIR = f"{DATA_DIR}/Onedrive"
EXPORT_DIR = os.environ.get("EXPORT_DIR", "EXPORT")

RENAME_MAP = {
    "VR / VE / Cuantía de la contratación": "Valor Referencial / Valor Estimado"
}

#
# Util
#

def recreate_folder(path):
    if os.path.exists(path):
        if os.path.isdir(path):
            shutil.rmtree(path)
        else:
            raise ValueError("Path exists but is not a directory")
    os.makedirs(path)

def validate_dataframe_header(df, rename_map=None):
    if not isinstance(df, pd.DataFrame):
        raise TypeError("Input 'df' must be a pandas DataFrame.")
    if rename_map is None:
        rename_map = {}
    if not isinstance(rename_map, dict):
        raise TypeError("Input 'rename_map' must be a dictionary or None.")

    required_header = [
        'N°',
        'Nombre o Sigla de la Entidad',
        'Fecha y Hora de Publicacion',
        'Nomenclatura',
        'Reiniciado Desde',
        'Objeto de Contratación',
        'Descripción de Objeto',
        'Valor Referencial / Valor Estimado',
        'Moneda',
        'Versión SEACE'
    ]

    current_header = df.columns.tolist()

    if current_header == required_header:
        print("Header is valid.")
        return df

    print("Initial header check failed. Attempting renames based on provided map...")

    if not rename_map:
        print("No rename map provided or map is empty.")
        raise ValueError(
            "Header is invalid and no rename map was provided to attempt corrections.\n"
            f"Expected: {required_header}\n"
            f"Got: {current_header}"
        )

    try:
        df_renamed = df.rename(columns=rename_map, errors='ignore')
        renamed_header = df_renamed.columns.tolist()

        # Check if any rename actually happened
        if renamed_header != current_header:
             print(f"Applied potential renames. Columns after attempting rename: {renamed_header}")
        else:
             print("No applicable columns found for renaming based on the provided map keys.")

        if renamed_header == required_header:
            print("Header is now valid after applying potential renames.")
            return df_renamed
        else:
            raise ValueError(
                "Header is invalid even after attempting renames from the map.\n"
                f"Expected: {required_header}\n"
                f"Got (after potential renames): {renamed_header}"
            )
    except Exception as e:
         raise ValueError(f"An error occurred during column renaming: {e}")

async def general_query_data_recursive(get_data, browser, year, start_date, end_date, opts):
    # Ensure the date range is valid
    if start_date > end_date:
        return pd.DataFrame()

    # Query data between start_date and end_date
    df = await get_data(browser, year, start_date, end_date, opts)
    
    # DEBUG
    print(f"MAIN: query_data_recursive: {start_date} {end_date} {len(df)}")

    if len(df) < LIMIT_QUERY or start_date == end_date:
        return df
    
    # Calculate a midpoint date within the range.
    delta_days = (end_date - start_date).days
    mid_date = start_date + datetime.timedelta(days=delta_days // 2)
    
    # To avoid potential infinite recursion if the split doesn't reduce the range,
    # make sure the midpoint is strictly before the end_date.
    if mid_date >= end_date:
        return df

    # Recursively query the two halves of the date range.
    left_df = await general_query_data_recursive(get_data, browser, year, start_date, mid_date, opts)
    right_df = await general_query_data_recursive(get_data, browser, year, mid_date + datetime.timedelta(days=1), end_date, opts)
    
    # Concatenate the results from the two halves.
    return pd.concat([left_df, right_df], ignore_index=True)

def prepare_data_for_excel(df_map, filter_filepath):
    if os.path.exists(filter_filepath):
        xls = pd.ExcelFile(filter_filepath)
        for key, df in df_map.items():
            sheet_name = key.capitalize()
            if sheet_name in xls.sheet_names:
                existing_df = pd.read_excel(xls, sheet_name)
                new_rows = df[~df['Nomenclatura'].isin(existing_df['Nomenclatura'])]
                df_map[key] = pd.concat([existing_df, new_rows], ignore_index=True)

def format_table(wb, sheetname, df, display_name, output_file, temp_index_array=None):
    # Define styles
    style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top', horizontal='left')
    right_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top', horizontal='right')
    font = openpyxl.styles.Font(name="Aptos Narrow", size=10)

    width_dict = {
        "Descripción de Objeto": 60,
        "Nombre o Sigla de la Entidad": 40,
        "Valor Referencial / Valor Estimado": 30
    }

    # Apply the styles
    table = openpyxl.worksheet.table.Table(
        displayName=display_name,
        ref=f'A1:{openpyxl.utils.get_column_letter(df.shape[1])}{len(df)+1}'
    )
    table.tableStyleInfo = style
    wb[sheetname].add_table(table)

    # Set column widths based on column names
    for idx, col_name in enumerate(df.columns, start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        width = width_dict.get(col_name, 25)
        wb[sheetname].column_dimensions[col_letter].width = width

    # Determine target column index for "Valor Referencial / Valor Estimado"
    try:
        target_index = list(df.columns).index("Valor Referencial / Valor Estimado") + 1
    except ValueError:
        target_index = None

    # Load the previous workbook if it exists and the sheetname is available
    if os.path.exists(output_file):
        old_wb = openpyxl.load_workbook(output_file)
        if sheetname in old_wb.sheetnames:
            old_sheet = old_wb[sheetname]
        else:
            old_sheet = None
    else:
        old_sheet = None

    # Apply alignment to all cells in the table
    for new_row_idx, row in enumerate(
            wb[sheetname].iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=df.shape[1]),
            start=1):
        if new_row_idx == 1:
            for cell in row:
                cell.font = font
                cell.alignment = alignment
            continue
        old_row_idx = temp_index_array[new_row_idx - 2] + 1 if temp_index_array is not None else new_row_idx
        for cell in row:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            # Copy fill style from the old workbook if the sheet exists and the cell has a fill style
            if old_sheet is not None:
                old_cell = old_sheet[f'{col_letter}{old_row_idx}']
                if old_cell.fill and old_cell.fill.fill_type is not None:
                    cell.fill = copy(old_cell.fill)
            # Apply font and alignment
            cell.font = font
            if target_index and cell.column == target_index:
                cell.alignment = right_alignment
            else:
                cell.alignment = alignment

    # Set row height to auto
    for row_num in range(2, len(df) + 2):
        wb[sheetname].row_dimensions[row_num].height = 90

def data_to_excel(keyword_dfs, output_file):
    filepath = f"{TMP_DIR}/{uuid.uuid4()}.xlsx"
    temp_indexes = {}
    # Export all DataFrames to an Excel file with each on a separate sheet.
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for keyword, df_kw in keyword_dfs.items():
            # Add a temporary index column to preserve the original row positions
            df_kw['temp_index'] = range(1, len(df_kw) + 1)

            # Sort the DataFrame by 'Fecha y Hora de Publicacion' in descending order
            df_kw['valor_datetime'] = pd.to_datetime(
                df_kw['Fecha y Hora de Publicacion'],
                format='%d/%m/%Y %H:%M'
            )
            df_kw.sort_values(by='valor_datetime', ascending=False, inplace=True)
            df_kw.drop('valor_datetime', axis=1, inplace=True)

            # Extract the temporary index as an array and then remove the column
            temp_indexes[keyword] = df_kw['temp_index'].to_numpy()
            df_kw.drop(columns=['temp_index'], inplace=True)

            df_kw.to_excel(writer, sheet_name=keyword.capitalize(), index=False)

    # Format table
    wb = openpyxl.load_workbook(filename = filepath)
    for keyword, df_kw in keyword_dfs.items():
        # Retrieve the temp index array for this sheet, if available
        temp_index_array = temp_indexes.get(keyword, None)

        format_table(wb, keyword.capitalize(), df_kw, keyword.replace(" ", "").capitalize(), output_file, temp_index_array)
    wb.save(output_file)

#
# OBRAS
#

async def get_data_obras(browser, year, start_date, end_date, opts):
    # Check if end_date is before start_date
    if end_date < start_date:
        raise ValueError("end_date cannot be before start_date")

    # Format and print the dates
    f_start_date = start_date.strftime("%d/%m/%Y")
    f_end_date = end_date.strftime("%d/%m/%Y")

    # Start page instance
    context = await browser.new_context()
    page = await context.new_page()

    # Web scraping
    await page.goto("https://prod2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml")
    await page.get_by_role("link", name="Buscador de Procedimientos de").click()
    time.sleep(9)
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:anioConvocatoria_label\"]").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:anioConvocatoria_panel\"]").get_by_text(year).click()
    await page.locator("[id^=\"tbBuscador\\:idFormBuscarProceso\\:j_idt\"][id$=\"_panel\"]").get_by_text("Obra", exact=True).dispatch_event("click")
    await page.get_by_text("Búsqueda Avanzada").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dfechaInicio_input\"]").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dfechaInicio_input\"]").fill(f_start_date)
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dfechaFin_input\"]").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dfechaFin_input\"]").fill(f_end_date)
    await page.get_by_role("button", name="Buscar").click()
    #await page.pause()

    # Wait for the filter
    #await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dtProcesos_data\"]").filter(has_not_text="No se encontraron Datos").click()
    time.sleep(6)

    # Download the results
    async with page.expect_download() as download_info:
        await page.get_by_role("button", name="Exportar a Excel").click()
    download = await download_info.value
    filepath = f"{TMP_DIR}/{uuid.uuid4()}.xls"
    await download.save_as(filepath)

    # Cleanup
    await page.close()
    await context.close()

    df = pd.read_excel(filepath)

    df = validate_dataframe_header(df, RENAME_MAP)

    return df

async def query_obras_data(browser, year, current_date):
    async def query_data_recursive(start_date, end_date):
        return await general_query_data_recursive(get_data_obras, browser, year, start_date, end_date, {})

    given_year = int(year)
    results = []
    
    # Only proceed if given_year is less than or equal to current year
    if given_year > current_date.year:
        return pd.DataFrame()
    
    # 1. Process the given year in 15-day chunks
    # If the given year is the current year, query only until current_date,
    # otherwise query the full year (January 1 to December 31)
    start_date_given = datetime.date(given_year, 1, 1)
    end_date_given = current_date if given_year == current_date.year else datetime.date(given_year, 12, 31)
    
    cur_date = start_date_given
    while cur_date <= end_date_given:
        # Define a 15-day window (including cur_date)
        next_date = cur_date + datetime.timedelta(days=14)
        if next_date > end_date_given:
            next_date = end_date_given
        df_part = await query_data_recursive(cur_date, next_date)
        results.append(df_part)
        # Move to the day after next_date for the next interval.
        cur_date = next_date + datetime.timedelta(days=1)
    
    # 2. Process each full year after the given year up to (but not including) the current year.
    for yr in range(given_year + 1, current_date.year):
        start_date_year = datetime.date(yr, 1, 1)
        end_date_year = datetime.date(yr, 12, 31)
        mid_date = start_date_year + (end_date_year - start_date_year) // 2
        df_first_half = await query_data_recursive(start_date_year, mid_date)
        df_second_half = await query_data_recursive(mid_date + datetime.timedelta(days=1), end_date_year)
        results.append(pd.concat([df_first_half, df_second_half], ignore_index=True))

    # 3. Process the current year (if it's after the given year) from January 1 to today.
    if current_date.year > given_year:
        start_date_current = datetime.date(current_date.year, 1, 1)
        end_date_current = current_date
        if (end_date_current - start_date_current).days + 1 > 300:
            mid_date = start_date_current + (end_date_current - start_date_current) // 2
            df_first_half = await query_data_recursive(start_date_current, mid_date)
            df_second_half = await query_data_recursive(mid_date + timedelta(days=1), end_date_current)
            df_current = pd.concat([df_first_half, df_second_half], ignore_index=True)
        else:
            df_current = await query_data_recursive(start_date_current, end_date_current)
        results.append(df_current)

    # Combine all data into one DataFrame
    if results:
        return pd.concat(results, ignore_index=True)
    else:
        return pd.DataFrame()

def filter_data_obras(df, lower_bound):
    def convert_to_float(x):
        try:
            return float(x.replace(',', ''))
        except Exception:
            return np.nan

    # Drop the "N°" column.
    if "N°" in df.columns:
        df = df.drop("N°", axis=1)

    # Create a helper numeric column for filtering and sorting.
    df['valor_numeric'] = df["Valor Referencial / Valor Estimado"].apply(convert_to_float)

    # Filter the DataFrame:
    mask = (df['valor_numeric'] > lower_bound) | (df['valor_numeric'].isna())
    df_filtered = df[mask].copy()

    # Sort the filtered DataFrame in descending order using the numeric column.
    df_sorted = df_filtered.sort_values(by='valor_numeric', ascending=False, na_position='first')

    # Optionally drop the helper column if no longer needed.
    df_sorted = df_sorted.drop('valor_numeric', axis=1)

    dfs = {}
    for keyword in KEYWORDS:
        dfs[keyword] = df_sorted[df_sorted["Descripción de Objeto"].str.contains(keyword, case=False, na=False)]
    dfs = {MAIN_SHEET_NAME: df_sorted, **dfs}

    return dfs

#
# VIDRIOS
#

async def get_data_vidrios(browser, year, start_date, end_date, opts):
    filtro = opts.get("filter")
    # Check if end_date is before start_date
    if end_date < start_date:
        raise ValueError("end_date cannot be before start_date")

    # Format and print the dates
    f_start_date = start_date.strftime("%d/%m/%Y")
    f_end_date = end_date.strftime("%d/%m/%Y")

    # Start page instance
    context = await browser.new_context()
    page = await context.new_page()

    # Web scraping
    await page.goto("https://prod2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml")
    await page.get_by_role("link", name="Buscador de Procedimientos de").click()
    time.sleep(9)
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:anioConvocatoria_label\"]").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:anioConvocatoria_panel\"]").get_by_text(year).click()
    await page.get_by_text("Búsqueda Avanzada").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dfechaInicio_input\"]").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dfechaInicio_input\"]").fill(f_start_date)
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dfechaFin_input\"]").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dfechaFin_input\"]").fill(f_end_date)
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:descripcionObjeto\"]").click()
    await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:descripcionObjeto\"]").fill(filtro)
    #await page.pause()
    await page.get_by_role("button", name="Buscar").click()

    # Wait for the filter
    #await page.locator("[id=\"tbBuscador\\:idFormBuscarProceso\\:dtProcesos_data\"]").filter(has_not_text="No se encontraron Datos").click()
    time.sleep(6)

    # Download the results
    async with page.expect_download() as download_info:
        await page.get_by_role("button", name="Exportar a Excel").click()
    download = await download_info.value
    filepath = f"{TMP_DIR}/{uuid.uuid4()}.xls"
    await download.save_as(filepath)

    # Cleanup
    await page.close()
    await context.close()

    df = pd.read_excel(filepath)
    df = df[::-1].reset_index(drop=True)

    df = validate_dataframe_header(df, RENAME_MAP)

    return df

async def query_vidrios_data(browser, year, current_date):
    async def query_data_recursive(filter, start_date, end_date):
        return await general_query_data_recursive(get_data_vidrios, browser, year, start_date, end_date, {"filter": filter})

    given_year = int(year)
    df_map = {}

    # Only proceed if given_year is less than or equal to current year
    if given_year > current_date.year:
        return pd.DataFrame()

    global_results = []
    for filter in KEYWORDS_VIDRIOS:
        results = []

        start_date_given = datetime.date(given_year, 1, 1)
        end_date_given = current_date if given_year == current_date.year else datetime.date(given_year, 12, 31)
        
        cur_date = start_date_given
        while cur_date <= end_date_given:
            next_date = cur_date + datetime.timedelta(days=300)
            if next_date > end_date_given:
                next_date = end_date_given
            df_part = await query_data_recursive(filter, cur_date, next_date)
            results.append(df_part)
            cur_date = next_date + datetime.timedelta(days=1)

        # Combine all data into one DataFrame
        df = pd.concat(results, ignore_index=True)[::-1].reset_index(drop=True)
        global_results.append(df)
    global_df = pd.concat(global_results, ignore_index=True)

    # Drop the "N°" column.
    if "N°" in global_df.columns:
        global_df = global_df.drop("N°", axis=1)

    df_map[MAIN_SHEET_NAME] = global_df

    return df_map

#
# Main
#

async def main():
    if not os.path.isdir(DATA_DIR):
        raise FileNotFoundError(f"Directory {DATA_DIR} does not exist!")

    recreate_folder(TMP_DIR)
    recreate_folder(DRIVE_DIR)
    os.makedirs(QUERY_DIR, exist_ok=True)
    os.makedirs(f"{DRIVE_DIR}/{EXPORT_DIR}", exist_ok=True)

    # Import data
    if os.environ.get("INCREMENTAL") != "no":
        result = subprocess.run([
            "onedrive",
            "--sync",
            "--syncdir",
            DRIVE_DIR,
            "--single-directory",
            EXPORT_DIR,
            "--download-only",
            "--cleanup-local-files",
        ])
        if result.returncode != 0:
            sys.exit(result.returncode)

    # Get date data
    timezone = pytz.timezone('America/Lima')
    now = datetime.datetime.now(timezone)
    current_date = now.date()

    async with async_playwright() as p:
        if os.environ.get("ENV") == "dev":
            browser = await p.chromium.launch(headless=False, args=['--ozone-platform=wayland'])
        else:
            browser = await p.chromium.launch(headless=True)

        # Vidrio fetch
        year = str(current_date.year)
        filter_filepath = f"{DRIVE_DIR}/{EXPORT_DIR}/SEACE_VIDRIOS_{year}.xlsx"
        df_map = await query_vidrios_data(browser, year, current_date)
        prepare_data_for_excel(df_map, filter_filepath)
        data_to_excel(df_map, filter_filepath)

        # Obras fetch
        for year in [str(current_date.year - i) for i in range(1)]:
            print(f"MAIN: Starting data collection for year {year}.")
            export_filepath = f"{QUERY_DIR}/{year}.xlsx"
            filter_filepath = f"{DRIVE_DIR}/{EXPORT_DIR}/SEACE_OBRAS_{year}.xlsx"
            if os.path.exists(export_filepath) and year != str(current_date.year):
                print(f"MAIN: {export_filepath} already exists, skipping query.")
                df = pd.read_excel(export_filepath)
            else:
                df = await query_obras_data(browser, year, current_date)
                df.to_excel(export_filepath, index=False)
            df_map = filter_data_obras(df, 4000000)
            prepare_data_for_excel(df_map, filter_filepath)
            data_to_excel(df_map, filter_filepath)

        # Cleanup
        await browser.close()

    # Export data
    result = subprocess.run([
        "onedrive",
        "--sync",
        "--syncdir",
        DRIVE_DIR,
        "--single-directory",
        EXPORT_DIR,
        "--upload-only",
        "--no-remote-delete",
    ])
    if result.returncode != 0:
        sys.exit(result.returncode)

asyncio.run(main())

